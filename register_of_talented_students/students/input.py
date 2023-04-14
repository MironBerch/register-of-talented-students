from openpyxl import load_workbook, Workbook

from students.services import (
    create_student, search_student, update_student, deduct_students,
)
from school_classes.services import (
    get_school_class, create_school_class_if_not_exist,
)


def import_students(filepath):
    """Import students excel table to database"""
    deduct_students()
    workbook = load_workbook(filename=filepath)

    sheet = workbook.worksheets[0]

    name_index = 4
    surname_index = 3
    patronymic_index = 5
    class_digit_index = 1
    class_letter_index = 2

    for row in sheet.iter_rows():

        if row[0].value == '№ п/п':
            continue

        name = str(row[name_index].value)
        surname = str(row[surname_index].value)
        patronymic = str(row[patronymic_index].value)
        class_digit = str(row[class_digit_index].value)
        class_letter = str(row[class_letter_index].value)

        school_class = class_digit + class_letter

        if not patronymic:
            patronymic = ' '

        full_name = surname + ' ' + name + ' ' + patronymic

        create_school_class_if_not_exist(school_class=school_class)

        if search_student(full_name=full_name):
            update_student(
                full_name=full_name,
                school_class=get_school_class(school_class),
                is_learns=True,
            )
        else:
            create_student(
                full_name=full_name,
                school_class=get_school_class(school_class),
                is_learns=True,
            )


def create_import_students_example() -> None:
    """Create example students excel table"""

    wb = Workbook()

    ws = wb.active

    ws.title = 'example'

    ws['A1'] = '№ п/п'
    ws['B1'] = 'Этап (числовое обозначение)'
    ws['C1'] = 'Наименование учебного коллектива'
    ws['D1'] = 'Фамилия'
    ws['E1'] = 'Имя'
    ws['F1'] = 'Отчество'

    wb.save('media//students_import_example.xlsx')
