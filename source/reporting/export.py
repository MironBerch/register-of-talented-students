from openpyxl import Workbook


def format(input_date: str) -> str:
    day = input_date[8:10]
    mounth = input_date[5:7]
    year = input_date[0:4]
    date = day + '.' + mounth + '.' + year
    return date

def export_contest_to_excel(contests) -> None:
    contests = contests.order_by('creation_date')

    wb = Workbook()

    ws = wb.active

    ws.title = "2023"

    ws['A1'] = 'Название мероприятия'
    ws['B1'] = 'ФИО ученика'
    ws['C1'] = 'ФИО преподавателя'
    ws['D1'] = 'Предмет'
    ws['E1'] = 'Класс'
    ws['F1'] = 'Направление'
    ws['G1'] = 'Этап'
    ws['H1'] = 'Результат'
    ws['I1'] = 'Комментарий'
    ws['J1'] = 'Дата мероприятия'
    ws['K1'] = 'Дата Учёта'

    letter_string = 'ABCDEFGHIJK'
    contest_index = 1

    for contest in contests:
        contest_index = contest_index + 1
        for index in range(11):
            cell = letter_string[index] + str(contest_index)
            if letter_string[index] == 'A':
                ws[cell] = str(contest.title)
            if letter_string[index] == 'B':
                ws[cell] = str(contest.students_name)
            if letter_string[index] == 'C':
                ws[cell] = str(contest.teachers_name)
            if letter_string[index] == 'D':
                ws[cell] = str(contest.subject)
            if letter_string[index] == 'E':
                ws[cell] = str(contest.school_сlass)
            if letter_string[index] == 'F':
                ws[cell] = str(contest.direction)
            if letter_string[index] == 'G':
                ws[cell] = str(contest.stage)
            if letter_string[index] == 'H':
                ws[cell] = str(contest.result)
            if letter_string[index] == 'I':
                ws[cell] = str(contest.other)
            if letter_string[index] == 'J':
                ws[cell] = format(input_date=str(contest.event_date))
            if letter_string[index] == 'K':
                ws[cell] = format(input_date=str(contest.creation_date))

    wb.save('media//учёт.xlsx')